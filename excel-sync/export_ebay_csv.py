"""Export available lots from D1 into an eBay bulk-listing prefill file.

Queries D1 directly via `wrangler d1 execute ... --json` (no hand-written
row data), then builds a brand-new, minimal single-sheet workbook and
writes one row per available lot starting at row 4.

Rows 1-3 reproduce eBay's own header structure -- the #INFO/Version/
Template markers, the merged "Set A" / "Set B" group labels, and the
five bordered, commented column headers -- by hand, rather than loading
and re-saving the original ebay_template.xlsx. That file is a real,
complex Excel export (embedded logo image, legacy VML cell comments, an
external-workbook link with Microsoft-specific extensions) that openpyxl
does not round-trip losslessly: re-saving it silently drops the image,
converts shared strings to inline strings, and regenerates the comments/
VML relationship structure with non-standard relationship IDs -- Excel
then flags the result as corrupt and "repairs" it on open. None of that
structure is needed for what this script does (it only ever touches the
"eBay-prefill-listing-template" sheet), so building fresh avoids the
round-trip entirely. If eBay changes their template format, re-extract
rows 1-3 from a fresh template download and update HEADER_* below.

ebay_template.xlsx itself is left untouched as a human-readable
reference copy of the real template -- it is not read by this script.
"""

import json
import os
import subprocess
import sys

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
EXPORT_PATH = os.path.join(SCRIPT_DIR, "ebay_export.xlsx")

SHEET_NAME = "eBay-prefill-listing-template"
FIRST_DATA_ROW = 4
TITLE_MAX_LEN = 80

D1_DATABASE = "holly-hill-shop"
D1_QUERY = "SELECT id, name, category, image_urls FROM lots WHERE status='available';"

# Hand-extracted from ebay_template.xlsx's "eBay-prefill-listing-template"
# sheet, rows 1-3 -- see the module docstring for why this is hardcoded
# rather than read from the file at runtime.
HEADER_COLUMN_WIDTHS = {"A": 18.33, "B": 16.66, "C": 41.16, "D": 23.66, "E": 44.5, "F": 10.66}
HEADER_COLUMNS = [
    (
        "A", "Custom Label (SKU)",
        "Optional column. Useful for tracking the output",
    ),
    (
        "B", "Item Photo URL",
        "At least 1 web-hosted image link for the item you are trying to list. "
        "Up to 24 images can be provided for each item.\n\nExample: URL 1 | URL 2 | URL 3",
    ),
    ("C", "Title", "Title for the item you are trying to list"),
    ("D", "Category", "Category you list with in other marketplaces or your own store"),
    (
        "E", "Aspects",
        "Specify aspects that you use on other marketplaces or your own store as "
        "pipe-separated name value pairs. E.g. Color=Red|Size=Small. This field is "
        "optional and can be left blank.",
    ),
]

# Plain RGB approximations of the original's theme-indexed colors -- exact
# theme-color fidelity isn't load-bearing to eBay's parser, and a fresh
# workbook has its own default theme, so reusing the original's theme
# index numbers wouldn't reliably reproduce the same colors anyway.
INFO_FILL = PatternFill("solid", fgColor="F2F2F2")
SET_A_FILL = PatternFill("solid", fgColor="DDEBF7")
SET_B_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
BOTTOM_THIN_BORDER = Border(bottom=Side(style="thin"))
TOP_ALIGN = Alignment(vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
BOLD = Font(bold=True)


class ExportError(Exception):
    pass


def fetch_available_lots():
    cmd = (
        f'npx wrangler d1 execute {D1_DATABASE} --remote '
        f'--command "{D1_QUERY}" --json'
    )
    try:
        proc = subprocess.run(
            cmd, cwd=REPO_ROOT, capture_output=True, text=True, timeout=60, shell=True
        )
    except subprocess.TimeoutExpired as err:
        raise ExportError(f"wrangler d1 execute timed out: {err}")

    if proc.returncode != 0:
        raise ExportError(f"wrangler d1 execute failed:\n{proc.stderr or proc.stdout}")

    try:
        data = json.loads(proc.stdout)
    except json.JSONDecodeError as err:
        raise ExportError(
            f"Could not parse wrangler's --json output: {err}\nOutput was:\n{proc.stdout}"
        )

    if not data or "results" not in data[0]:
        raise ExportError(f"Unexpected wrangler --json shape: {data}")

    return data[0]["results"]


def parse_image_urls(raw):
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return []
    return parsed if isinstance(parsed, list) else []


def build_minimal_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Row 1 -- #INFO metadata
    ws["A1"] = "#INFO"
    ws["B1"] = "Version=1.0.0"
    ws["D1"] = "Template=eBay-taxonomy-mapping-template_US"
    for coord in ("A1", "B1", "D1"):
        ws[coord].number_format = "@"
        ws[coord].fill = INFO_FILL
    ws["A1"].alignment = TOP_ALIGN

    # Row 2 -- #INFO + merged "Set A" / "Set B" group labels
    ws["A2"] = "#INFO"
    ws["A2"].number_format = "@"
    ws["A2"].fill = INFO_FILL
    ws["A2"].alignment = TOP_ALIGN

    ws["B2"] = "Set A"
    ws["D2"] = "Set B"
    ws.merge_cells("B2:C2")
    ws.merge_cells("D2:E2")
    for coord in ("B2", "D2"):
        ws[coord].font = BOLD
        ws[coord].number_format = "@"
        ws[coord].alignment = CENTER_ALIGN
        ws[coord].border = BOTTOM_THIN_BORDER
    ws["B2"].fill = SET_A_FILL
    ws["D2"].fill = SET_B_FILL
    ws.row_dimensions[2].height = 34

    # Row 3 -- the actual column headers, bordered and commented
    for col, label, comment_text in HEADER_COLUMNS:
        cell = ws[f"{col}3"]
        cell.value = label
        cell.font = BOLD
        cell.number_format = "@"
        cell.border = THIN_BORDER
        cell.fill = HEADER_FILL
        cell.comment = Comment(comment_text, "eBay")

    for col, width in HEADER_COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    return wb


def main():
    lots = fetch_available_lots()

    wb = build_minimal_workbook()
    ws = wb[SHEET_NAME]

    no_photo_lots = []
    row = FIRST_DATA_ROW
    for lot in lots:
        lot_id = lot.get("id") or ""
        name = lot.get("name") or ""
        category = lot.get("category") or None
        image_urls = parse_image_urls(lot.get("image_urls"))

        title = name if len(name) <= TITLE_MAX_LEN else name[:TITLE_MAX_LEN]
        photo_url_field = "|".join(image_urls)

        sku_cell = ws.cell(row=row, column=1, value=lot_id)  # Custom Label (SKU)
        sku_cell.number_format = "@"  # keep numeric-looking SKUs as text
        ws.cell(row=row, column=2, value=photo_url_field)    # Item Photo URL
        ws.cell(row=row, column=3, value=title)              # Title
        ws.cell(row=row, column=4, value=category)           # Category
        # Column E (Aspects) intentionally left blank.

        if not image_urls:
            no_photo_lots.append(lot_id)

        row += 1

    wb.save(EXPORT_PATH)

    print(f"OK: exported {len(lots)} lot(s) to {EXPORT_PATH}")
    if no_photo_lots:
        print(
            f"WARNING: {len(no_photo_lots)} lot(s) have NO photos "
            "(weak/no eBay AI suggestions): " + ", ".join(no_photo_lots)
        )


if __name__ == "__main__":
    try:
        main()
    except ExportError as err:
        print(f"FAILED: {err}")
        sys.exit(1)
    except Exception as err:
        print(f"FAILED: Unexpected error: {err}")
        sys.exit(1)
